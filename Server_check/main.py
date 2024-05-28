from openpyxl import Workbook, load_workbook


"""
LUU Y: DONG FILE XLSX TRUOC KHI CHAY
"""

# chuan hoa string
def standardize(string):
    res = string
    while res[0] == ' ':
        res = res[1:]
    while res[-1] == ' ':
        res = res[:1]
    return res


# doc tu file list ip
def read_data_from_compare():
    wb = load_workbook('ServerCompare.xlsx')
    sheet = wb.active  # Assuming data is on the first sheet

    # Start from B2 and iterate until empty row is found
    live_ip_temp = {}
    for row in sheet.iter_rows(min_row=2):  # Skip the first row (header)
        cell_a = row[0]  # Access cell B (second column)
        cell_b = row[1]  # Access cell C (third column)

        # Check if one of cells are empty (consider different data types like None, empty strings)
        if cell_a.value is None or cell_b.value is None:
            break  # Stop iterating if both B and C are empty
        live_ip_temp[cell_b.value] = cell_a.value
    return live_ip_temp


live_ip = read_data_from_compare()


# kiem tra cac ip trong tung o cua ServerNeedCheck/ luu gia tri ip va name
def ip_check(ip_list):
    ip = ""
    name = ""
    while ip_list != "":
        ip_list = standardize(ip_list)
        x = ip_list.partition(',')[0]
        ip_list = ip_list.partition(',')[2]
        x = standardize(x)
        temp = live_ip.get(x, 0)
        if temp != 0:
            ip = x
            name = temp
            break
    return ip, name

def read_data_from_check():
    wb = load_workbook('ServerNeedCheck.xlsx')  # Load the workbook
    sheet = wb.active # Access the specified sheet

    for row in sheet.iter_rows(min_row=3):  # Iterate from start_row
        cell_d = row[3]  # Access cell D (fourth column)
        cell_g = row[6]  # Target cell G (seventh column)
        cell_h = row[7]  # Target cell H (eighth column)
        cell_i = row[8]
        l = str(cell_d.value)
        ip_wr, name_wr = ip_check(l)
        if ip_wr != "" and name_wr != "":
            cell_g.value = ip_wr  # Write data from B to G
            cell_h.value = name_wr  # Write data from D to H
            cell_i.value = "MATCHED"
    wb.save('ServerNeedCheck.xlsx')  # Save changes to the workbook
    return print("Changed file")


read_data_from_check()
